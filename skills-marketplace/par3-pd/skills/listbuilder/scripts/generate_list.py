#!/usr/bin/env python3
"""
Generic prospect list generator. Reads a JSON file with dynamic columns,
triggers, and rows. Outputs a formatted Excel file + CSV.

Usage:
    python generate_list.py --input data.json --output-dir ./output --filename "my-list" --hot-threshold 15 --warm-threshold 10
"""
import json, csv, argparse, os, sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1B2631")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Arial")
DATA_FONT = Font(size=10, name="Arial")
LINK_FONT = Font(size=10, name="Arial", color="0563C1", underline="single")
HOT_FILL = PatternFill("solid", fgColor="D5F5E3")
WARM_FILL = PatternFill("solid", fgColor="FEF9E7")
COLD_FILL = PatternFill("solid", fgColor="FADBD8")
THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def is_missing(val):
    if val is None:
        return True
    s = str(val).strip().lower()
    return s in ("", "unknown", "not found", "n/a", "none")


def build_prospect_sheet(wb, rows, columns, hot_thresh, warm_thresh):
    ws = wb.active
    ws.title = "Prospect List"
    ws.freeze_panes = "A2"

    for col_idx, col_def in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def.get("width", 20)

    for row_idx, data in enumerate(rows, 2):
        score = 0
        try:
            score = int(data.get("prospect_score", 0))
        except (ValueError, TypeError):
            pass

        if score >= hot_thresh:
            row_fill = HOT_FILL
        elif score >= warm_thresh:
            row_fill = WARM_FILL
        else:
            row_fill = COLD_FILL

        for col_idx, col_def in enumerate(columns, 1):
            field = col_def["field"]
            col_type = col_def.get("type", "text")
            val = data.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_type == "url" and val and not is_missing(val):
                cell.value = str(val)
                try:
                    cell.hyperlink = str(val)
                except Exception:
                    pass
                cell.font = LINK_FONT
            elif col_type in ("number", "score") and val and not is_missing(val):
                try:
                    cell.value = int(val)
                except (ValueError, TypeError):
                    try:
                        cell.value = float(val)
                    except (ValueError, TypeError):
                        cell.value = str(val)
                cell.font = DATA_FONT
                cell.alignment = RIGHT
            else:
                cell.value = str(val) if val else ""
                cell.font = DATA_FONT

            cell.border = THIN_BORDER
            cell.fill = row_fill
            if col_type not in ("number", "score"):
                cell.alignment = LEFT

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"


def build_summary_sheet(wb, rows, hot_thresh, warm_thresh):
    ws = wb.create_sheet("Summary")
    title_font = Font(bold=True, size=14, name="Arial")
    label_font = Font(bold=True, size=11, name="Arial")
    value_font = Font(size=11, name="Arial")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25

    ws.cell(row=1, column=1, value="Prospect List Summary").font = title_font
    row_num = 3

    companies = set(r.get("company_name", "") for r in rows)
    scores = []
    for r in rows:
        try:
            scores.append(int(r.get("prospect_score", 0)))
        except (ValueError, TypeError):
            pass

    hot = sum(1 for s in scores if s >= hot_thresh)
    warm = sum(1 for s in scores if warm_thresh <= s < hot_thresh)
    cold = sum(1 for s in scores if s < warm_thresh)

    stats = [
        ("Total Contacts (rows)", len(rows)),
        ("Unique Companies", len(companies)),
        ("", ""),
        (f"HOT ({hot_thresh}+)", hot),
        (f"WARM ({warm_thresh}-{hot_thresh - 1})", warm),
        (f"COLD (<{warm_thresh})", cold),
        ("Avg Score", round(sum(scores) / len(scores), 1) if scores else 0),
        ("", ""),
        ("With LinkedIn", sum(1 for r in rows if not is_missing(r.get("person_linkedin")))),
        ("With Email", sum(1 for r in rows if not is_missing(r.get("person_email")))),
    ]

    for label, val in stats:
        if label == "":
            row_num += 1
            continue
        ws.cell(row=row_num, column=1, value=label).font = label_font
        ws.cell(row=row_num, column=2, value=val).font = value_font
        ws.cell(row=row_num, column=1).border = THIN_BORDER
        ws.cell(row=row_num, column=2).border = THIN_BORDER
        row_num += 1

    row_num += 1
    ws.cell(row=row_num, column=1, value="By Role").font = title_font
    row_num += 1
    role_counts = {}
    for r in rows:
        role = r.get("role", "Unknown")
        role_counts[role] = role_counts.get(role, 0) + 1
    for role, count in sorted(role_counts.items(), key=lambda x: -x[1]):
        ws.cell(row=row_num, column=1, value=role).font = label_font
        ws.cell(row=row_num, column=2, value=count).font = value_font
        ws.cell(row=row_num, column=1).border = THIN_BORDER
        ws.cell(row=row_num, column=2).border = THIN_BORDER
        row_num += 1


def build_scoring_sheet(wb, triggers, hot_thresh, warm_thresh):
    ws = wb.create_sheet("Scoring Guide")
    title_font = Font(bold=True, size=14, name="Arial")
    header_font = Font(bold=True, size=11, name="Arial", color="FFFFFF")
    label_font = Font(bold=True, size=11, name="Arial")
    value_font = Font(size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="1B2631")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 60

    ws.cell(row=1, column=1, value="Prospect Scoring Guide").font = title_font

    row_num = 3
    for col, header in [(1, "Trigger"), (2, "Points"), (3, "Why This Matters")]:
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = THIN_BORDER

    max_score = 0
    for t in triggers:
        row_num += 1
        pts = t.get("points", 0)
        max_score += pts
        ws.cell(row=row_num, column=1, value=t.get("name", "")).font = label_font
        ws.cell(row=row_num, column=2, value=pts).font = value_font
        ws.cell(row=row_num, column=3, value=t.get("why", "")).font = value_font
        for c in range(1, 4):
            ws.cell(row=row_num, column=c).border = THIN_BORDER

    row_num += 1
    ws.cell(row=row_num, column=1, value="MAX SCORE").font = Font(bold=True, size=12, name="Arial")
    ws.cell(row=row_num, column=2, value=max_score).font = Font(bold=True, size=12, name="Arial")

    row_num += 2
    ws.cell(row=row_num, column=1, value="Score Ranges:").font = title_font
    row_num += 1
    ranges = [
        (f"HOT ({hot_thresh}+)", "Green rows — highest priority prospects", HOT_FILL),
        (f"WARM ({warm_thresh}-{hot_thresh - 1})", "Yellow rows — good potential, worth reaching out", WARM_FILL),
        (f"COLD (<{warm_thresh})", "Red rows — lower priority but still in ICP", COLD_FILL),
    ]
    for label, desc, fill in ranges:
        ws.cell(row=row_num, column=1, value=label).font = label_font
        ws.cell(row=row_num, column=1).fill = fill
        ws.cell(row=row_num, column=2).fill = fill
        ws.cell(row=row_num, column=3, value=desc).font = value_font
        ws.cell(row=row_num, column=3).fill = fill
        row_num += 1


def write_csv(rows, columns, filepath):
    fields = [c["field"] for c in columns]
    headers = [c["header"] for c in columns]
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in rows:
            writer.writerow([str(r.get(field, "")) for field in fields])


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--filename", default="prospect-list")
    parser.add_argument("--hot-threshold", type=int, default=15)
    parser.add_argument("--warm-threshold", type=int, default=10)
    args = parser.parse_args()

    with open(args.input) as f:
        data = json.load(f)

    rows = data.get("rows", [])
    triggers = data.get("triggers", [])
    columns = data.get("columns", [])

    # If no columns provided, use defaults
    if not columns:
        columns = [
            {"header": "Company Name", "field": "company_name", "width": 26, "type": "text"},
            {"header": "Website", "field": "website", "width": 28, "type": "url"},
            {"header": "Industry", "field": "industry", "width": 20, "type": "text"},
            {"header": "HQ Location", "field": "hq_location", "width": 20, "type": "text"},
            {"header": "Employee Count", "field": "employee_count", "width": 14, "type": "number"},
            {"header": "Prospect Score", "field": "prospect_score", "width": 16, "type": "score"},
            {"header": "Score Breakdown", "field": "score_breakdown", "width": 40, "type": "text"},
            {"header": "First Name", "field": "person_first_name", "width": 16, "type": "text"},
            {"header": "Last Name", "field": "person_last_name", "width": 16, "type": "text"},
            {"header": "Role", "field": "role", "width": 22, "type": "text"},
            {"header": "LinkedIn", "field": "person_linkedin", "width": 36, "type": "url"},
            {"header": "Email", "field": "person_email", "width": 28, "type": "text"},
        ]

    # Sort by prospect_score descending
    def sort_key(r):
        try:
            return -int(r.get("prospect_score", 0))
        except (ValueError, TypeError):
            return 0
    rows.sort(key=sort_key)

    os.makedirs(args.output_dir, exist_ok=True)

    wb = Workbook()
    build_prospect_sheet(wb, rows, columns, args.hot_threshold, args.warm_threshold)
    build_summary_sheet(wb, rows, args.hot_threshold, args.warm_threshold)
    if triggers:
        build_scoring_sheet(wb, triggers, args.hot_threshold, args.warm_threshold)

    xlsx_path = os.path.join(args.output_dir, f"{args.filename}.xlsx")
    csv_path = os.path.join(args.output_dir, f"{args.filename}.csv")

    wb.save(xlsx_path)
    write_csv(rows, columns, csv_path)

    scores = []
    for r in rows:
        try:
            scores.append(int(r.get("prospect_score", 0)))
        except (ValueError, TypeError):
            pass

    print(json.dumps({
        "status": "success",
        "xlsx_path": xlsx_path,
        "csv_path": csv_path,
        "total_contacts": len(rows),
        "unique_companies": len(set(r.get("company_name", "") for r in rows)),
        "hot": sum(1 for s in scores if s >= args.hot_threshold),
        "warm": sum(1 for s in scores if args.warm_threshold <= s < args.hot_threshold),
        "cold": sum(1 for s in scores if s < args.warm_threshold),
    }, indent=2))


if __name__ == "__main__":
    main()
